import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from util import get_molar_masses, formula_parser, mass_calculator
import os

def process_formulas(formulas, filename, total_masses=[0.10, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50]):  # change total mass here if needed
    molar_masses = get_molar_masses.read_molar_masses(filename)
    wb = Workbook()
    
    # Define the light yellow fill
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    for total_mass in total_masses:
        ws = wb.create_sheet(title=f"{total_mass:.2f}")
        max_elements = max(len(formula_parser.get_parsed_formula(formula)) for formula in formulas)
        headers = ['Formula', 'Total Mass']
        for i in range(max_elements):
            headers.extend([f'Element{i+1}', f'Mass{i+1}'])
        ws.append(headers)
        
        for formula in formulas:
            ratios = formula_parser.get_parsed_formula(formula)
            masses = mass_calculator.calculate_masses(ratios, total_mass, molar_masses)
            row = [formula, total_mass]
            for element, mass in masses.items():
                row.extend([element, f"{mass:.4f}"])  # Format to 4 decimal places
            ws.append(row)
        
        # Apply the yellow fill to the 'Mass' columns
        for col in range(4, 4 + 2 * max_elements, 2):  # 'Mass' columns are 4th, 6th, 8th, etc.
            for row in range(2, len(formulas) + 2):  # Skip the header row, and apply to data rows
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.remove(wb['Sheet'])
    # Determine a unique filename
    base_filename = 'calculated'
    extension = '.xlsx'
    n = 1
    while os.path.exists(f"{base_filename}_{n}{extension}"):
        n += 1
    unique_filename = f"{base_filename}_{n}{extension}"
    
    wb.save(unique_filename)
    print(f"File saved as {unique_filename}")

def read_formulas_from_text():
    formulas = []
    print("Paste the formulas (one per line), and press 'Enter' on an empty line when finished:")
    while True:
        line = input()
        if line == '':
            break
        formulas.append(line.strip())
    return formulas

def read_formulas_from_txt_file(filepath):
    with open(filepath, 'r') as file:
        formulas = [line.strip() for line in file.readlines()]
    return formulas

def read_formulas_from_excel(filepath):
    df = pd.read_excel(filepath, header=None)
    return df[0].tolist()
